#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Created on Sat Aug 04 20:14:11 2018

@author: Ghost
"""
import pandas as pd
import numpy as np
import objects
import math
import random
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
import copy
class GetOutOfLoop( Exception ):
    pass


def getDemandList(route, sheet):
    df = pd.read_excel(route,  sheetname=sheet)
    data = df.values.tolist()
    #data = transpuesta(data)
    return data

def letJustDepartmentDemand(demand):
    #IN: Type: List of Demand
    bufferDic = {}
    for i in range(len(demand)):
        try:
            if(str(demand[i][0])[0:2] == u"IE" or str(demand[i][0])[0:2] == u"IM"):
                actualDemandDic = { demand[i][0] : int(math.ceil(demand[i][4])) }
                bufferDic.update(actualDemandDic)
        except UnicodeEncodeError:
            None
    #OUT: type: DIC - { Code : student_Demand }
    return bufferDic

def print_r(matriz):
    for fila in matriz:
        print fila

def transpuesta(matriz):
    rows = len(matriz)
    cols = len(matriz[0])
    return [[matriz[j][i] for j in xrange(rows)] for i in xrange(cols)]

def getDataList(route):
    df = pd.read_csv(route)
    data = df.values.tolist()
    data = transpuesta(data)
    return data

def createListOfObjects(dataPlan, pensum):
    listOfObjects = []
    for index in range(len(dataPlan[0])):

        newObject = objects.course(dataPlan[13][index], dataPlan[15][index], dataPlan[16][index], str(dataPlan[17][index])[0], str(dataPlan[17][index])[1], str(dataPlan[17][index])[2], dataPlan[20][index], pensum, 0, 0)
        listOfObjects.append(newObject)

    #OUT: List of Objects "Course"
    return listOfObjects

def compare(L1, L2):
    if len(L1) <> len(L2):
        return None
    else:
        differ = []
        for pos in range(len(L1)):
            if L1[pos] <> L2[pos]:
                differ.append(pos)
                return differ

def mergeAllCourses(dataPlanList, demand, maxTheory, maxLeftOver):
    objectCoursesList = []
    for i in range(len(dataPlanList)):
        objectCoursesList.append(createListOfObjects(dataPlanList[i], i))

    for i in range(len(objectCoursesList)-1):
        objectCoursesList[0].extend(objectCoursesList[i+1])


    for e in range(len(objectCoursesList[0])):
        for i in range(len(objectCoursesList[0])):
            try:
                if(objectCoursesList[0][e].code == objectCoursesList[0][i].code and
                   e != i):
                    objectCoursesList[0][i] = None
            except AttributeError:
                None
                #print( str(counter)+" -- "+str(i))

    departmentList = []
    exitData = {}
    for i in range(len(objectCoursesList[0])):
        try:
            objectCoursesList[0][i].code
            if(str(objectCoursesList[0][i].code) == 'nan'):
                None

            if(str(objectCoursesList[0][i].code)[0:2] == 'IE' or
               str(objectCoursesList[0][i].code)[0:2] == 'MT'):
                departmentList.append(objectCoursesList[0][i])
        except AttributeError:
            None

    for i in range(len(departmentList)):
        #Llenando demanda; y cantidad de secciones que necesitara el curso en funcion del numero de maximo para teoria y demanda de estudiantes
        if(departmentList[i].code in demand):
            departmentList[i].demand = demand[departmentList[i].code]
            departmentList[i].sections = departmentList[i].demand/maxTheory
            if(departmentList[i].demand % maxTheory > maxLeftOver):
                departmentList[i].sections = departmentList[i].sections + 1
            #print departmentList[i].sections
            exitData.update({ departmentList[i].code : departmentList[i] })
            #print departmentList[i].theoryPeriods

    return exitData

def verifyTeacherAv(listOfTeachers, listOfCourses):
    exitData = []
    error = None
    listOfTeachers2 = listOfTeachers
    listOfTeachers = []
    for key in listOfTeachers2:
        listOfTeachers.append(listOfTeachers2[key])

    for t in range(len(listOfTeachers)):
        #print listOfTeachers[t].labPeriods
    #for t in range(len(listOfTeachers)):
        totalPeriods = 0
        for i in range(len(listOfTeachers[t].workTime)):
            #print listOfTeachers[t].workTime
            for j in range(len(listOfTeachers[t].workTime[i])):
                #print listOfTeachers[t].workTime[i][j]
                if(listOfTeachers[t].workTime[i][j] == "x"):
                    #print listOfTeachers[t].workTime[i][j]
                    totalPeriods = totalPeriods+1
                    #print "sumando"
                    None
        periodsCount = 0
        for key in listOfTeachers[t].theoryPeriods:
            try:
                periodsCount = periodsCount + int(listOfCourses[key].theoryPeriods)*int(listOfTeachers[t].theoryPeriods[key])
            except KeyError:
                error = "ERROR: El profesor "+listOfTeachers[t].name+" tiene asignado un curso que no corresponde a la demanda"
                break

        for key in listOfTeachers[t].labPeriods:
            try:
                periodsCount = periodsCount + int(listOfCourses[key].labPeriods)*int(listOfTeachers[t].labPeriods[key])
            except KeyError:
                error = "ERROR: El profesor "+listOfTeachers[t].name+" tiene asignado un curso que no corresponde a la demanda"
                break
        #print str(periodsCount)+" -  "+str(totalPeriods)
        if(periodsCount > totalPeriods):
            error = "ERROR: El profesor "+listOfTeachers[t].name+u" Tiene asignados más periodos que espacio de trabajo"

    #print listOfTeachers[0].workTime
        exitBuffer = { listOfTeachers[t].name : [periodsCount, totalPeriods] }
        exitData.append(exitBuffer)
    #print error
    return [exitData , error ]

def readTeachers(route,allCourses):
    xls = pd.ExcelFile(route)
    teacherList = {}

    for indexProfesor in range(len(xls.sheet_names)):
        actualTeacher = pd.read_excel(route, xls.sheet_names[indexProfesor])
        actualTeacher = actualTeacher.values.tolist()
        coursesList = {}
        demandLeft = {}
        name = actualTeacher[0][4].replace(" ", "")+"_"+actualTeacher[1][4].replace(" ", "")
        workTime = actualTeacher[3:21]
        workMatrix = []
        theoryPeriods = {}
        #usedTime = []
        #print workTime
        labPeriods = {}
        prettyWorkTime = [] #CAMBIAR SI ES NECESARIO CONVERTIRLO A DIC
        for i in range(len(workTime)):
            if(i == 4):
                None
            else:
                workMatrix.append(workTime[i][3:8])

            if(str(workTime[i][8]) != 'nan'):
                coursesList.update({ workTime[i][8]: {"TH" : workTime[i][9] , "LAB": workTime[i][10]} } )
                #print allCourses[workTime[i][8]].labPeriods
                demandLeft.update({ workTime[i][8]: {"TH" : workTime[i][9]*int(allCourses[workTime[i][8]].theoryPeriods), "LAB": workTime[i][10]*int(allCourses[workTime[i][8]].labPeriods) } } )
                bufferDic = { workTime[i][8]:workTime[i][9] }
                theoryPeriods.update(bufferDic)
                bufferDic = { workTime[i][8]:workTime[i][10] }
                labPeriods.update(bufferDic)

            for teacherWorkPeriod in range(3,8,1):
                if( workTime[i][teacherWorkPeriod] == 'x'):
                    if(i < 4):
                        prettyWorkTime.append([i,teacherWorkPeriod-3])
                    if(i > 4):
                        prettyWorkTime.append([i-1,teacherWorkPeriod-3])

        #print name
        #print workMatrix
        #print len(prettyWorkTime)
        #print name
        #print demandLeft
        teacherList.update({ name : objects.teacher( name, workMatrix, theoryPeriods, labPeriods, coursesList, prettyWorkTime, [], demandLeft)})
    #data = transpuesta(data)

    #print teacherData[0][4]
    #print teacherList
    return teacherList

def loadForbiddens(route):
    xls = pd.ExcelFile(route)
    forbiddenList = []
    for indexYear in range(len(xls.sheet_names)):
        actualYear = pd.read_excel(route, xls.sheet_names[indexYear])

        actualYear = actualYear.values.tolist()
        workTime = actualYear[3:21]
        workMatrix = []
        #print workTime
        prettyForbiddenTime = []
        for i in range(len(workTime)):
            if(i == 4):
                None
            else:
                workMatrix.append(workTime[i][3:8])

            for teacherWorkPeriod in range(3,8,1):
                if( workTime[i][teacherWorkPeriod] == 'x'):
                    if(i < 4):
                        prettyForbiddenTime.append([i,teacherWorkPeriod-3])
                    if(i > 4):
                        prettyForbiddenTime.append([i-1,teacherWorkPeriod-3])


        forbiddenList.append(objects.forbiddenTime(xls.sheet_names[indexYear], workMatrix, prettyForbiddenTime))
        #print objects.forbiddenTime(xls.sheet_names[indexYear], workMatrix)
        #forbiddenList.append(objects.forbiddenTime( year, forbiddenPeriods))
    #data = transpuesta(data)

    #print teacherData[0][4]
    return forbiddenList


def InitialSolution(allCourses, labMax, labLeftOver, nLabs):
    labSections = 0
    allSections = []
    contadorSecciones = 0
    contadorSeccioensLab = 0

    for key in allCourses:
        #print key+"  "+str(allCourses[key].demand)
        for s in range(allCourses[key].sections):
            sectionObject = objects.section(allCourses[key], None, None, 'TH', str(s+1)+'0'+key)
            sectionObject.labNumber = ''
            allSections.append(sectionObject)
            #print  str(s+1)+'0'+key+' TH'
            #print key+"  "+'TH'
            contadorSecciones = contadorSecciones + 1

            #print str(s+1)+'0'+key

        if(int(allCourses[key].labPeriods) != 0):
            #print allCourses[key].labPeriods

            if(allCourses[key].demand % labMax > labLeftOver):
                labSections = allCourses[key].demand/labMax + 1
            else:
                labSections = allCourses[key].demand/labMax

            #print labSections
            labSection = 1
            contador = 1
            for l in range(labSections):
                if(contador == 2):
                    sectionObject = objects.section(allCourses[key], None, None, 'LAB', str(labSection)+str(contador)+key)
                    sectionObject.labNumber = random.randrange(0,nLabs,1)
                    allSections.append(sectionObject)
                    #print ""
                    #print str(labSection)+str(contador)+key+' LAB's
                    #print key+"  "+'LAB'
                    contadorSeccioensLab = contadorSeccioensLab +1
                    contador = 1
                    labSection = labSection +1

                else:
                    sectionObject =(objects.section(allCourses[key], None, None, 'LAB', str(labSection)+str(contador)+key))
                    sectionObject.labNumber = random.randrange(0,nLabs,1)
                    allSections.append(sectionObject)
                    #print str(labSection)+str(contador)+key+' LAB'
                    contador = contador + 1



    return allSections


def random_assign(allPeriods, allTeachers, forbiddenTime, allLabs):
    generation = [] #Salida Generaciones Iniciales Asignadas
    FTIME = {}
    codeMemory = {}
    labMemory = {}

    for P in range(len(allPeriods)): # Se guarda un registro a que periodo pertenece cual maestro
        codeMemory.update( { allPeriods[P][allPeriods[P].keys()[0]].code : None })
        labMemory.update( { allPeriods[P][allPeriods[P].keys()[0]].code : None })

    #print codeMemory
    # Se creará un dic de ForbiddenTime por simplicidad
    for F in forbiddenTime:
        FTIME.update( { int(F.year[-1:]) : F })
    #print FTIME

    for P in range(len(allPeriods)):
        #print P

        #Valores Generales
        periodCode = allPeriods[P][allPeriods[P].keys()[0]].code
        courseCode = allPeriods[P][allPeriods[P].keys()[0]].course.code
        theoryPeriods = allPeriods[P][allPeriods[P].keys()[0]].course.theoryPeriods
        labPeriods = allPeriods[P][allPeriods[P].keys()[0]].course.labPeriods
        yearPeriod = allPeriods[P][allPeriods[P].keys()[0]].course.year
        classType = allPeriods[P][allPeriods[P].keys()[0]].classType

        #Valores que deben ser específicos
        while True:
            randomPeriod = random.randrange(0,17,1)
            randomDay = random.randrange(0,5,1)

            randomTeacherPosition = random.randrange(0,len(allTeachers),1)
            randomTeacher = allTeachers[allTeachers.keys()[randomTeacherPosition]]
            #print labMemory[periodCode]
            if(labMemory[periodCode] == None):
                randomLabNumber = random.randrange(0,len(allLabs),1)
            else:
                randomLabNumber = labMemory[periodCode]
                #print periodCode+" - "+str(randomLabNumber)
            randomLab = allLabs[randomLabNumber]
            #print randomLab

            # ----------------- Condiciones para  conseguir un Periodo,Día,Profesor Perfecto -------------------

            # ---------- ForbiddenTime ---------------
            forbiddenTimeCondition = [randomPeriod, randomDay] not in FTIME[yearPeriod].prettyForbiddenTime

            #  --------------- Dia y Periodo no ocupado para el profesor ---------------
            periodForTeacherCondition = [randomPeriod, randomDay]  not in randomTeacher.usedTime

            # ------ La clase la puede dar el profesor -------------

            classForTeacherCondition = courseCode in randomTeacher.coursesList

            # ------- Demand disponible -----------

            if(classType == 'LAB' and classForTeacherCondition):
                teacherDemandCondition = randomTeacher.demandLeft[courseCode]['LAB'] > 0
            elif(classType == 'TH' and classForTeacherCondition):
                teacherDemandCondition = randomTeacher.demandLeft[courseCode]['TH'] > 0

            #  ---------------Lab Conditions  ---------------
            if(classType == 'LAB'):
                labConditions = [randomPeriod, randomDay] not in randomLab.prettyUsedTime
            else:
                labConditions = True #Al no ser un laboratorio entonces pasa

            # ------ Seccion Sin dueño -------------

            if(codeMemory[periodCode] == None or
               codeMemory[periodCode] == randomTeacher.name):

                   teacherFromTheSectionCondition = True
            else:
                teacherFromTheSectionCondition = False

            # Execute conditions
            if(forbiddenTimeCondition and
               periodForTeacherCondition and
               labConditions and
               classForTeacherCondition and
               teacherDemandCondition and
               teacherFromTheSectionCondition):
                   #print [randomPeriod, randomDay]
                   break


        #Assignation space

        #Lab Assignation Space
        if(classType == 'LAB'):

            randomLab.prettyUsedTime.append([randomPeriod, randomDay])
            allLabs[randomLab.number] = randomLab #Actualizar Labs
            labMemory[periodCode] = randomLabNumber # Agregando lab a la sección
            allPeriods[P][allPeriods[P].keys()[0]].labNumber = randomLabNumber

            if(randomTeacher.demandLeft[courseCode]['LAB'] > 0):  #Se resta un periodo de la demanda
                randomTeacher.demandLeft[courseCode]['LAB'] = randomTeacher.demandLeft[courseCode]['LAB'] - 1
        else:
            #TH Assignation Space
            if(randomTeacher.demandLeft[courseCode]['TH'] > 0):
                randomTeacher.demandLeft[courseCode]['TH'] = randomTeacher.demandLeft[courseCode]['TH'] - 1



        #General Assignation Space
        randomTeacher.usedTime.append([randomPeriod, randomDay]) #Se Agrega periodo utilizado al teacher
        codeMemory[periodCode] = randomTeacher.name #Agregando maestro a la sección



        #randomTeacher.sections.update

        allTeachers[randomTeacher.name] = randomTeacher #Actualizar Teachers

#        print " ------------------------- "
#        #print P
#        print [randomPeriod,randomDay]
#        print periodCode
#        print "---------------------------"
        newPeriod = objects.period(
                    allPeriods[P][allPeriods[P].keys()[0]],
                    periodCode,
                    randomPeriod,
                    randomDay,
                    randomTeacher.name)

        generation.append(newPeriod)







        #print periodDay

    # TEST COURSES
#    for codes in codeMemory:
#        print codes+" - "+codeMemory[codes]
    # TEST LABS
    #print "I---------------------"
    for keys in allLabs:
        #print keys
        #print allLabs[keys].prettyUsedTime
        for element in allLabs[keys].prettyUsedTime:
            if(allLabs[keys].prettyUsedTime.count(element) > 1):
                print "REPETIDO EN LABS!!!"


    # TEST TEACHERS
    for keys in allTeachers:
        #print keys
        #print allTeachers[keys].demandLeft
        for elemento in allTeachers[keys].usedTime:
            #print "REPETIDOS: "
            if(allTeachers[keys].usedTime.count(elemento) > 1):
                print "REPETIDO EN TEACHERS!!!!!"

    labsCheck = { 0 : [] , 1: []}
    for i in range(len(generation)):
        if(generation[i].section.classType == 'LAB'):
            labsCheck[generation[i].section.labNumber].append( [generation[i].period, generation[i].day] )
    #print labsCheck
    for keys in labsCheck:
        #print keys
        #print allLabs[keys].prettyUsedTime
        for i in range(len(labsCheck[keys])):
            if( labsCheck[keys][i] not in allLabs[keys].prettyUsedTime or
                allLabs[keys].prettyUsedTime[i] not in labsCheck[keys] ):
                    print "ERROR"

    #print "F----------------------------"

#    for P in range(len(generation)):
#        if(generation[P].section.classType == 'LAB'):
#            if( ([generation[P].period , generation[P].day] in allLabs[generation[P].section.labNumber].prettyUsedTime) == False):
#                print generation[P].code



    return generation, allPeriods, allTeachers, forbiddenTime, allLabs


def FITNESS1(solution, allTeachers, allLabs, nlabs, forbiddenTime):
################ CONTROL POINTS ########################################

    closeLab = 40  #Puntos por laboratorio cercano en el horario
    closeTH = 50    #Puntos por Periodo de teoría cercano en el horario
    sameDay = 20    #Puntos por que los periodos se encuentren el mismo día
    differentDay_SameTime = 1  #Puntos por que los periodos se encuentren en diferente día pero al mismo periodo


    score = 0

##############################################################

#    for P in range(len(solution)):
#        if(solution[P].section.classType == 'LAB'):
#            if([solution[P].period, solution[P].day] not in allLabs[solution[P].section.labNumber].prettyUsedTime):
#                print solution[P].code

    # PUNTOS 1 : Labs Juntos
    for periods in range(len(solution)):
        #print allLabs
        if(solution[periods].section.classType == 'LAB'):  #PUNTOS TIPO LAB
            if(solution[periods].code == allLabs[solution[periods].section.labNumber].time[solution[periods].period][solution[periods].day]  ):

                if(solution[periods].period > 0 and solution[periods].period < 16):
                    if(solution[periods].code == allLabs[solution[periods].section.labNumber].time[solution[periods].period - 1][solution[periods].day]):

                        score = score + closeLab

                    if(solution[periods].code == allLabs[solution[periods].section.labNumber].time[solution[periods].period + 1][solution[periods].day]):

                        score = score + closeLab
                else:
                    if(solution[periods].period == 0):
                        if(solution[periods].code == allLabs[solution[periods].section.labNumber].time[solution[periods].period + 1][solution[periods].day]):
                            score = score + closeLab

                        if(solution[periods].code == allLabs[solution[periods].section.labNumber].time[solution[periods].period + 2][solution[periods].day]):
                            score = score + closeLab
                    elif(solution[periods].period == 16):
                        if(solution[periods].code == allLabs[solution[periods].section.labNumber].time[solution[periods].period - 1][solution[periods].day]):
                            score = score + closeLab

                        if(solution[periods].code == allLabs[solution[periods].section.labNumber].time[solution[periods].period - 2][solution[periods].day]):
                            score = score + closeLab

    # PUNTOS 3: Periodos Teoría Juntos
        if(solution[periods].section.classType == 'TH'):
            for reCheck in range(len(solution)):
                if(solution[periods].day == solution[reCheck].day):
                    score = score + sameDay
                else:
                    if(solution[periods].period == solution[reCheck].period):
                        score = differentDay_SameTime
                if(solution[periods].period == solution[reCheck].period - 1):
                    score = score + closeTH
                if(solution[periods].period == solution[reCheck].period - 2):
                   score = score + closeTH
                if(solution[periods].period == solution[reCheck].period + 1):
                   score = score + closeTH
                if(solution[periods].period == solution[reCheck].period + 2 ):
                   score = score + closeTH



    return score


def Crossover(generation, allTeachers, allLabs, labs, forbiddenTime):
    # ----- For Mutation ----

    # ----------------
    # ---------     CROSSOVER TIPO #1: Uniforme: --------------
    totalCrossed = 0



    for P in range(len(generation)):


        #print P
        doIT = random.randrange(0,2,1)  #doIT LAB - Move Day and period

        if(doIT == 1):
            totalCrossed = totalCrossed + 1
            classType = generation[P].section.classType

            # CROSS:  Actual Vars  <----> Random Vars FOR LAB
            #Actal Vars
            actualPeriod = generation[P].period
            actualDay = generation[P].day
            actualTeacher = allTeachers[generation[P].teacher]
            actualCode = generation[P].section.course.code
            if(classType == 'LAB'):
                actualLab = generation[P].section.labNumber
                actualSectionCode = generation[P].section.code
            while True:


                #Random Vars
                while True:
                    print 'manolo'
                    randomPeriodObject = random.randrange(0,len(generation),1)
                    randomPeriod = generation[randomPeriodObject].period
                    randomDay = generation[randomPeriodObject].day
                    randomTeacher = allTeachers[generation[randomPeriodObject].teacher]
                    #print randomTeacher.name
                    #print P
                    #print generation[randomPeriodObject].section.classType == classType
                    #print [randomPeriod, randomDay] not in actualTeacher.usedTime
                    #print [actualPeriod,actualDay] not in randomTeacher.usedTime
                    #print actualTeacher.usedTime
                    #print ""

                    #print actualTeacher.name
                    #print actualTeacher.usedTime
                    #print [randomPeriod,randomDay]
                    #print generation[randomPeriodObject].section.classType == classType
                    #print "PERIOD IN ACTUAL: "+str([randomPeriod,randomDay] not in actualTeacher.usedTime)
                    #print "PERIOD IN RANDOM: "+str( [actualPeriod,actualDay] not in randomTeacher.usedTime)
                    if(generation[randomPeriodObject].section.classType == classType and
                       [randomPeriod,randomDay] not in actualTeacher.usedTime and
                       [actualPeriod,actualDay] not in randomTeacher.usedTime):

                        #print "JEJE"
                        #print ""
                        #print  generation[randomPeriodObject].code+" - "+generation[P].code


                        #print randomTeacher.usedTime
                        randomCode = generation[randomPeriodObject].section.course.code

                        if(classType == 'LAB'):
                            randomLab = generation[randomPeriodObject].section.labNumber
                            randomSectionCode = generation[randomPeriodObject].section.code

                        break

#                if(P < 1):
#                    print ""
#                    print randomTeacher.name
#                    print len(randomTeacher.usedTime)
#                print ""
#                print randomTeacher.name
#                print randomTeacher.coursesList
#                print actualCode
                #print "ACTUAL_CODE IN RANDOM: "+str(actualCode in randomTeacher.coursesList)
                #print "RANDOM_CODE IN ACTUAL: "+str(randomCode in actualTeacher.coursesList)
                #print randomTeacher.name


                #print "LEVEL 1"
                if([randomPeriod,randomDay] in actualTeacher.prettyWorkTime and #Periodo en espacio de profesor
                   [actualPeriod,actualDay] in randomTeacher.prettyWorkTime):
                   #print "LEVEL 2"

                   #Crossing
                   generation[P].period = randomPeriod
                   generation[P].day = randomDay

                   generation[randomPeriodObject].period = actualPeriod
                   generation[randomPeriodObject].day = actualDay

#                   print "---------------------"
#                   print actualTeacher.name
#                   print [actualPeriod,actualDay]
#                   print actualTeacher.usedTime
#                   print "---------------------"

                   actualTeacher.usedTime.remove([actualPeriod,actualDay])
                   randomTeacher.usedTime.remove([randomPeriod,randomDay])

                   actualTeacher.usedTime.append([randomPeriod,randomDay])
                   randomTeacher.usedTime.append([actualPeriod,actualDay])

#                       try:
#                           actualTeacher.usedTime.remove([actualPeriod,actualDay])
#
#                       except ValueError:
#
#                           None
#                       actualTeacher.usedTime.append([randomPeriod,randomDay])
#
#                       try:
#                           randomTeacher.usedTime.remove([randomPeriod,randomDay])
#                       except ValueError:
#                           None
#
#                       randomTeacher.usedTime.append([actualPeriod,actualDay])
#
#                       allTeachers[generation[P].teacher] = actualTeacher
#                       allTeachers[generation[randomPeriodObject].teacher] = randomTeacher


                   if(classType == 'Lab'):
                       allLabs[randomLab].time[randomPeriod][randomDay] = actualSectionCode
                       allLabs[actualLab].time[actualPeriod][actualDay] = randomSectionCode

                       if([randomPeriod, randomDay] in allLabs[randomLab].prettyUsedTime):
                           allLabs[randomLab].prettyUsedTime.remove([randomPeriod, randomDay])
                       allLabs[randomLab].prettyUsedTime.append([actualPeriod,actualDay])

                       if([actualPeriod,actualDay] in allLabs[actualLab].prettyUsedTime):
                           allLabs[actualLab].prettyUsedTime.remove([actualPeriod,actualDay])
                       allLabs[actualLab].prettyUsedTime.append([randomPeriod, randomDay])



                   break


    #print totalCrossed

    #update usedTime



    #print ""
    #print position
#    teachersCheck = {}
#    usedMemory = {}
#    for key in allTeachers:
#        teachersCheck.update( {allTeachers[key].name : {} })
#        usedMemory.update( {allTeachers[key].name : [] })
#        #print allTeachers[position][key].usedTime
#
#    #print teachersCheck
#
#
#    for i in range(len(generation)):
#       if( generation[i].code in teachersCheck[generation[i].teacher]):
#           teachersCheck[generation[i].teacher].update( {generation[i].code : teachersCheck[generation[i].teacher][generation[i].code] + 1 } )
#       else:
#           teachersCheck[generation[i].teacher].update( {generation[i].code : 1 } )
#
#       usedMemory[generation[i].teacher].append([generation[i].period, generation[i].day])
#
#    for key in teachersCheck:
#        allTeachers[key].usedTime = teachersCheck[key]
#        allTeachers[key].usedMemory = usedMemory[key]

        #print allTeachers[key].usedTime
        else:
            #print "SALTO"
            None
    #print usedMemory



    return generation, allTeachers, allLabs, forbiddenTime, totalCrossed

def mutation(generation, allTeachers, allLabs, forbiddenTime, mutations):
    FTIME = {}
    for F in forbiddenTime:
        FTIME.update( { int(F.year[-1:]) : F })
    # ---------     CROSSOVER TIPO #1: Uniforme: --------------
    for P in range(mutations):
        # RANDOM PERIOD FOR MUTATION
        actualMutationObject = random.randrange(0,len(generation),1)
        classType = generation[actualMutationObject].section.classType
        actualPeriod = generation[actualMutationObject].period
        actualDay = generation[actualMutationObject].day
        actualTeacher = allTeachers[generation[actualMutationObject].teacher]
        actualCode = generation[actualMutationObject].code
        actualYear = generation[actualMutationObject].section.course.year
        actualLabNumber = generation[actualMutationObject].section.labNumber

        #print [actualPeriod, actualDay]
        #print actualTeacher.usedTime
        while True:
            randomPeriod = random.randrange(0,17,1)
            randomDay = random.randrange(0,5,1)

            #-------------- conditions to mutation -------------------

            #El profesor debe tener el periodo en su espacio de trabajo
            teacherForPeriodCondition = [randomPeriod, randomDay] in actualTeacher.prettyWorkTime

            #El profesor no debe tener asignado ese periodo
            timeNotUsedCondition = [randomPeriod, randomDay] not in actualTeacher.usedTime

            #El nuevo Periodo no debe estar en un horario prohibido de su año
            forbiddenCondition = [randomPeriod, randomDay] not in FTIME[actualYear].prettyForbiddenTime

            # ----- LABS Mutation Conditions
            if(classType == 'LAB'):
                labTimeCondition = [randomPeriod, randomDay] not in allLabs[actualLabNumber].prettyUsedTime
            else:
                labTimeCondition = True
            #print labTimeCondition
            # Ejecutando condiciones
            if(teacherForPeriodCondition and
               timeNotUsedCondition and
               forbiddenCondition and
               labTimeCondition):
                   None
                   break


#        print actualTeacher.name
#        print [actualPeriod, actualDay]
#        print [actualPeriod, actualDay] in allTeachers[actualTeacher.name].usedTime
#        #print actualTeacher.usedTime
#        print allLabs[actualLabNumber].prettyUsedTime
#        print actualLabNumber
#

#
#        if(classType == 'LAB'):
#            if(([actualPeriod, actualDay] in allLabs[actualLabNumber].prettyUsedTime) == False):
#                print actualCode
#                print "JEJE"


#            allLabs[actualLabNumber].prettyUsedTime.remove([actualPeriod, actualDay])
#            allLabs[actualLabNumber].prettyUsedTime.append([randomPeriod, randomDay])
#
#
#        allTeachers[actualTeacher.name].usedTime.remove([actualPeriod, actualDay])
#        allTeachers[actualTeacher.name].usedTime.append([randomPeriod, randomDay])
#
#        generation[P].period = randomPeriod
#        generation[P].day = randomDay









    return generation, allTeachers, allLabs, forbiddenTime

def testAssign(generation, allTeachers, forbiddenTime, allLabs):
    print " ------------------------- "
    	# 0 teachers
    # 1 labs
    show = 0
    # TEST COURSES
#    for codes in codeMemory:
#        print codes+" - "+codeMemory[codes]
    # TEST LABS
    #print "I---------------------"

    if(show == 0):
        for keys in allLabs:
            #print keys
            #print allLabs[keys].prettyUsedTime
            for element in allLabs[keys].prettyUsedTime:
                if(allLabs[keys].prettyUsedTime.count(element) > 1):
                    print "REPETIDO EN LABS!!!"


        # TEST TEACHERS
        teachersCheck = {}
        for keys in allTeachers:
            print keys
            #print allTeachers[keys].demandLeft
            print allTeachers[keys].usedTime
            print ""
            teachersCheck.update({ keys : []})
            for elemento in allTeachers[keys].usedTime:
                #print "REPETIDOS: "
                if(allTeachers[keys].usedTime.count(elemento) > 1):
                    print "REPETIDO EN TEACHERS!!!!!"

        labsCheck = { 0 : [] , 1: []}
        for i in range(len(generation)):
            if(generation[i].section.classType == 'LAB'):
                labsCheck[generation[i].section.labNumber].append( [generation[i].period, generation[i].day] )

            teachersCheck[generation[i].teacher].append([generation[i].period, generation[i].day])

        print teachersCheck
        #print labsCheck
        for keys in labsCheck:
            #print keys
            #print allLabs[keys].prettyUsedTime
            None
#            for i in range(len(labsCheck[keys])):
#                if( labsCheck[keys][i] not in allLabs[keys].prettyUsedTime or
#                    allLabs[keys].prettyUsedTime[i] not in labsCheck[keys] ):
#                        print "ERROR"

    if(show == 1):
        for keys in allLabs:
            print keys
            print allLabs[keys].prettyUsedTime
            print ""
            for element in allLabs[keys].prettyUsedTime:
                if(allLabs[keys].prettyUsedTime.count(element) > 1):
                    print "REPETIDO EN LABS!!!"


        # TEST TEACHERS
        teachersCheck = {}
        for keys in allTeachers:
            #print keys
            #print allTeachers[keys].demandLeft
            #print allTeachers[keys].usedTime
            #print ""
            teachersCheck.update({ keys : []})
            for elemento in allTeachers[keys].usedTime:
                #print "REPETIDOS: "
                if(allTeachers[keys].usedTime.count(elemento) > 1):
                    print "REPETIDO EN TEACHERS!!!!!"

        labsCheck = { 0 : [] , 1: []}
        for i in range(len(generation)):
            if(generation[i].section.classType == 'LAB'):
                labsCheck[generation[i].section.labNumber].append( [generation[i].period, generation[i].day] )

            teachersCheck[generation[i].teacher].append([generation[i].period, generation[i].day])

        #print teachersCheck
        #print labsCheck
        for keys in labsCheck:
            print keys
            print allLabs[keys].prettyUsedTime
            for i in range(len(labsCheck[keys])):
#                if( labsCheck[keys][i] not in allLabs[keys].prettyUsedTime or
#                    allLabs[keys].prettyUsedTime[i] not in labsCheck[keys] ):
#                        print "ERROR"
                None

    print "----------------------------"
    return

def selectAndReproduce(generation, allPeriods, allTeachers, forbiddenTime, allLabs, scores, maxNumber, minNumber):
    poblation = len(generation)

    new_generation = []
    new_scores = []
    new_allTeachers = []
    new_forbiddenTime = []
    new_allLabs = []

    for individual in range(maxNumber):
        ind = np.argmax(scores)
        #Se guardan las mejores soluciones
        new_generation.append(generation[ind])
        new_scores.append(scores[ind])
        new_allTeachers.append(allTeachers[ind])
        new_forbiddenTime.append(forbiddenTime[ind])
        new_allLabs.append(allLabs[ind])

    for individual in range(minNumber):
        ind = np.argmin(scores)
        #Se guardan las mejores soluciones
        new_generation.append(generation[ind])
        new_scores.append(scores[ind])
        new_allTeachers.append(allTeachers[ind])
        new_forbiddenTime.append(forbiddenTime[ind])
        new_allLabs.append(allLabs[ind])

    for individual in range( poblation - maxNumber - minNumber):
        ind = random.randrange(0,len(new_generation),1)

        new_generation.append(new_generation[ind])
        new_scores.append(new_scores[ind])
        new_allTeachers.append(new_allTeachers[ind])
        new_forbiddenTime.append(new_forbiddenTime[ind])
        new_allLabs.append(new_allLabs[ind])

    generation = new_generation
    scores = new_scores
    allTeachers = new_allTeachers
    forbiddenTime = new_forbiddenTime
    allLabs = new_allLabs


    return (generation, allPeriods, allTeachers, forbiddenTime, allLabs)

def teacherPeriods(generation):
# PROFESORES

    teacherShow = {}
    error = []
    for position in range(len(generation)):
        for pi in range(len(generation[position])):

            if(generation[position][pi].teacher in teacherShow ):
                #print pi
                if([str(generation[position][pi].day)+"-"+str(generation[position][pi].period)] in teacherShow[generation[position][pi].teacher]):
                    error.append("CRUCE PERIODOS DE PROFESOR: "+str(generation[position][pi].day)+"-"+str(generation[position][pi].period)+" en "+generation[position][pi].teacher)


                else:
                    # teacherShow with the section Code
                    #teacherShow[generation[position][pi].teacher].append([str(generation[position][pi].day)+"-"+str(generation[position][pi].period)+"-"+generation[position][pi].code])
                    # teacherShow without section Code
                    teacherShow[generation[position][pi].teacher].append([str(generation[position][pi].day)+"-"+str(generation[position][pi].period)])

                #periodShow.update( {generation[position][pi].code : periodShow[generation[position][pi].code].append([str(generation[position][pi].day)+"-"+str(generation[position][pi].period)])})
                #print periodShow[generation[position][pi].code]
            else:

                # teacher show with section Code
                #teacherShow.update( {generation[position][pi].teacher : [[str(generation[position][pi].day)+"-"+str(generation[position][pi].period)+"-"+generation[position][pi].code]] } )
                # teacher show without section code
                teacherShow.update( {generation[position][pi].teacher : [[str(generation[position][pi].day)+"-"+str(generation[position][pi].period)]] } )

            #print pi
    return teacherShow, error


def printDataTeacher(allTeachers):

    for key in allTeachers:
        print allTeachers[key].name
        print allTeachers[key].usedMemory
        print allTeachers[key].demandAssigned

def printDataPeriodsForCourse(generation):
    data = {}
    for i in range(len(generation)):
        periodCode = generation[i].section.course.code
        if(periodCode not in data):
            data.update( { periodCode : 1})
        else:
            data.update( { periodCode : data[periodCode] + 1})
    print data

def printDataSections(generation):
    data = {}
    for i in range(len(generation)):
        sectionCode = generation[i].section.code
        if(sectionCode not in data):
            data.update( { sectionCode : 1})
        else:
            data.update( { sectionCode : data[sectionCode] + 1})
    print data

def showCompleteData(generation):
    completeData = {}
    #print generation
    for P in range(len(generation)):
        if(int(generation[P].code[1]) == 0):
            #print generation[P].code
            completeData.update( {
                                    generation[P].code :  { 'LAB' : {} ,

                                                             'TH' : {
                                                                    'periodos' : [] , 'teacher': None, 'year': None
                                                                    }
                                                           }
                                  })

    for P in range(len(generation)):

        if(int(generation[P].code[1]) == 0):
            mainTheoryCode = generation[P].code
            #print mainTheoryCode
            bufferVar = completeData[mainTheoryCode]['TH']['periodos']
            bufferVar.append([generation[P].period , generation[P].day])
            completeData[mainTheoryCode]['TH']['periodos'] = bufferVar

            bufferVar = completeData[mainTheoryCode]['TH']['teacher']

            if(bufferVar == None):
                completeData[mainTheoryCode]['TH']['teacher'] = generation[P].teacher

            else:
                if(bufferVar != generation[P].teacher):
                    #print "ERROR -"+bufferVar+" - "+generation[P].teacher
                    None

            bufferVar = completeData[mainTheoryCode]['TH']['year']
            if(bufferVar == None):
                completeData[mainTheoryCode]['TH']['year'] = generation[P].section.course.year

            else:
                if(bufferVar != generation[P].section.course.year):
                    #print "ERROR -"+bufferVar+" - "+generation[P].teacher
                    None

    for P in range(len(generation)):

        if(int(generation[P].code[1]) != 0):
            #print 'jeje'
            #print generation[P].code
            #print ""
            #print completeData
            mainTheoryCode = generation[P].code[0]+"0"+generation[P].code[2:]
            #print mainTheoryCode
            if( generation[P].code not in completeData[mainTheoryCode]['LAB']):
                bufferDic = { generation[P].code : {} }
                bufferDic[generation[P].code].update( { 'periodos' : [ [generation[P].period, generation[P].day] ] ,
                                                        'teacher' : generation[P].teacher ,
                                                        'numLab' : generation[P].section.labNumber,
                                                        'year' : generation[P].section.course.year})

                completeData[mainTheoryCode]['LAB'].update(bufferDic)
            else:

                mainLabCode = generation[P].code
                periodsDic = completeData[mainTheoryCode]['LAB'][mainLabCode]['periodos']
                periodsDic.append([generation[P].period, generation[P].day])
                completeData[mainTheoryCode]['LAB'][mainLabCode]['periodos'] = periodsDic



    return completeData

def logger(individual, bestIndividual):
    show = False #MOSTRAR GENERACION NUEVA
    if(bestIndividual != ''):
        if( individual['score'] > bestIndividual['score']):
            bestIndividual = copy.deepcopy(individual)
            if(show == True):
                print "NEW MAX SCORE: "+str(bestIndividual['score'] )
                print showCompleteData(bestIndividual['individual'])
                print "---------------------------------------"
    else:
        bestIndividual = individual
        if(show == True):
            print "NEW MAX SCORE: "+str(bestIndividual['score'] )
            print showCompleteData(bestIndividual['individual'])
            print "---------------------------------------"
    return bestIndividual

def exitDataToExcel(bestIndividual, allLabs, forbiddenTime, allTeachers, writeLabs, writeTeachers, writeYear):

    redFill = PatternFill(start_color='FFFF0000',
                       end_color='FFFF0000',
                       fill_type='solid')

    yellowFill = PatternFill(start_color='E6FF18',
                       end_color='E6FF18',
                       fill_type='solid')

    whiteFill = PatternFill(start_color='FFFFFF',
                       end_color='FFFFFF',
                       fill_type='solid')
    completeData = showCompleteData(bestIndividual)

    if(writeLabs == True):
        # ESCRIBIENDO EXCEL DE LABORATORIOS -------------------------
        filepath="data/horarioLaboratorios.xlsx"
        wb = load_workbook(filepath)
        #Limpiando Hojas anteriores:
        for sheet in wb.sheetnames:
            if(sheet != 'Laboratorio 0'):
                wb.remove(wb[sheet])

        source = wb['Laboratorio 0']
        for keys in allLabs:
            if(keys != 0):
                bufferSheet = wb.copy_worksheet(source)
                bufferSheet.title = 'Laboratorio '+str(keys)


        for sheet in wb:
            #Preparing Data
            labNumber = sheet.title.split(' ')[1]
            sheet['E2'] = int(labNumber)
            cellToClean = ['D','E','F','G','H','I','J','K','L']
            for celda in range(len(cellToClean)):
                for i in range(18):
                    cell = cellToClean[celda]+str(i+5)
                    sheet[cell] = ''
                    sheet[cell].fill = whiteFill


            #Start writing data
            cell = 'I5'
            cellTeacher = 'J5'
            cellPeriod = ''
            for code in completeData:
                for labCode in completeData[code]['LAB']:
                    if(int(completeData[code]['LAB'][labCode]['numLab']) == int(labNumber)):

                        sheet[cell] = labCode
                        sheet[cellTeacher] = completeData[code]['LAB'][labCode]['teacher']
                        cell = cell[0]+str(int(cell[1:])+1)
                        cellTeacher = cellTeacher[0]+str(int(cellTeacher[1:])+1)
                        if(int(cell[1:])+1 == 24):
                            cell = 'K5'
                            cellTeacher = 'L5'

                        periodos = completeData[code]['LAB'][labCode]['periodos']
                        #print "PERIODOS:"
                        #print periodos
                        for P in range(len(periodos)):
                            period = periodos[P][0]
                            if(period >= 4):
                                period = period +1
                            day = periodos[P][1]
                            if(day == 0):
                                cellPeriod = 'D'
                            elif(day == 1):
                                cellPeriod = 'E'
                            elif(day == 2):
                                cellPeriod = 'F'
                            elif(day == 3):
                                cellPeriod = 'G'
                            elif(day == 4):
                                cellPeriod = 'H'
                            cellPeriod = cellPeriod+str(period+5)
                            bufferCell = sheet[cellPeriod].value
                            if(bufferCell == ''):
                                sheet[cellPeriod] = labCode
                                sheet[cellPeriod].font = Font(bold=True)
                            else:
                                sheet[cellPeriod].fill = redFill
                                sheet[cellPeriod].font = Font(bold=True)
                                sheet[cellPeriod] = bufferCell+", "+labCode




        wb.save(filepath)

    if(writeYear == True):

        # ESCRIBIENDO EXCEL DE LABORATORIOS -------------------------
        filepath="data/horarioEstudiantes.xlsx"
        wb = load_workbook(filepath)
        #Limpiando Hojas anteriores:
        for sheet in wb.sheetnames:
            if(sheet != 'Year 1'):
                wb.remove(wb[sheet])

        source = wb['Year 1']
        for keys in range(5):
            if(keys != 0):
                bufferSheet = wb.copy_worksheet(source)
                bufferSheet.title = 'Year '+str(keys + 1)


        for sheet in wb:
            #Preparing Data
            yearNumber = sheet.title.split(' ')[1]
            sheet['E2'] = int(yearNumber)
            cellToClean = ['D','E','F','G','H','I','J','K','L']
            for celda in range(len(cellToClean)):
                for i in range(18):
                    cell = cellToClean[celda]+str(i+5)
                    sheet[cell] = ''
                    sheet[cell].fill = whiteFill



            #Start writing data
            cell = 'I5'
            cellTeacher = 'J5'
            cellPeriod = ''
            for code in completeData:
                if(int(completeData[code]['TH']['year']) == int(yearNumber)):

                    sheet[cell] = code
                    sheet[cellTeacher] = completeData[code]['TH']['teacher']
                    cell = cell[0]+str(int(cell[1:])+1)
                    cellTeacher = cellTeacher[0]+str(int(cellTeacher[1:])+1)
                    if(int(cell[1:])+1 == 24):
                        cell = 'K5'
                        cellTeacher = 'L5'

                    periodos = completeData[code]['TH']['periodos']
                    for P in range(len(periodos)):
                        period = periodos[P][0]
                        if(period >= 4):
                            period = period +1
                        day = periodos[P][1]
                        if(day == 0):
                            cellPeriod = 'D'
                        elif(day == 1):
                            cellPeriod = 'E'
                        elif(day == 2):
                            cellPeriod = 'F'
                        elif(day == 3):
                            cellPeriod = 'G'
                        elif(day == 4):
                            cellPeriod = 'H'
                        cellPeriod = cellPeriod+str(period+5)
                        bufferCell = sheet[cellPeriod].value
                        if(bufferCell == ''):
                            sheet[cellPeriod] = code
                            sheet[cellPeriod].font = Font(bold=True)
                        else:
                            sheet[cellPeriod].fill = redFill
                            sheet[cellPeriod].font = Font(bold=True)
                            sheet[cellPeriod] = bufferCell+", "+code

            #Coloring forbiddenTime
            #print "RANGE: "
            #print len(forbiddenTime[int(yearNumber)-1].prettyForbiddenTime)
            for forbidden in range(len(forbiddenTime[int(yearNumber)-1].prettyForbiddenTime)):
                period = forbiddenTime[int(yearNumber)-1].prettyForbiddenTime[forbidden][0]
                if(period >= 4):
                    period = period +1
                day = forbiddenTime[int(yearNumber)-1].prettyForbiddenTime[forbidden][1]
                print[period,day]
                if(day == 0):
                    cellPeriod = 'D'
                elif(day == 1):
                    cellPeriod = 'E'
                elif(day == 2):
                    cellPeriod = 'F'
                elif(day == 3):
                    cellPeriod = 'G'
                elif(day == 4):
                    cellPeriod = 'H'
                cellPeriod = cellPeriod+str(period+5)
                sheet[cellPeriod].font = Font(bold=True)
                sheet[cellPeriod].fill = yellowFill


        wb.save(filepath)

    if(writeTeachers == True):
        # ESCRIBIENDO EXCEL DE LABORATORIOS -------------------------
        filepath="data/horarioProfesores.xlsx"
        wb = load_workbook(filepath)
        #Limpiando Hojas anteriores:
        for sheet in wb.sheetnames:
            if(sheet != 'Plantilla'):
                wb.remove(wb[sheet])

        source = wb['Plantilla']
        for keys in allTeachers:
            if(keys != 0):
                bufferSheet = wb.copy_worksheet(source)
                bufferSheet.title = keys[0:20]


        for sheet in wb:
            #Preparing Data
            teacherName = sheet.title
            sheet['E2'] = teacherName
            cellToClean = ['D','E','F','G','H','I','J','K','L']
            for celda in range(len(cellToClean)):
                for i in range(18):
                    cell = cellToClean[celda]+str(i+5)
                    sheet[cell] = ''
                    sheet[cell].fill = whiteFill


            #Start writing data
            cell = 'I5'
            cellTeacher = 'J5'
            cellPeriod = ''
            for code in completeData:
                for labCode in completeData[code]['LAB']:
                    actualTeacher = completeData[code]['LAB'][labCode]['teacher'][0:20]
                    if(teacherName == actualTeacher):

                        sheet[cell] = labCode
                        sheet[cellTeacher] = completeData[code]['LAB'][labCode]['teacher']
                        cell = cell[0]+str(int(cell[1:])+1)
                        cellTeacher = cellTeacher[0]+str(int(cellTeacher[1:])+1)
                        if(int(cell[1:])+1 == 24):
                            cell = 'K5'
                            cellTeacher = 'L5'

                        periodos = completeData[code]['LAB'][labCode]['periodos']
                        #print "PERIODOS:"
                        #print periodos
                        for P in range(len(periodos)):
                            period = periodos[P][0]
                            if(period >= 4):
                                period = period +1
                            day = periodos[P][1]
                            if(day == 0):
                                cellPeriod = 'D'
                            elif(day == 1):
                                cellPeriod = 'E'
                            elif(day == 2):
                                cellPeriod = 'F'
                            elif(day == 3):
                                cellPeriod = 'G'
                            elif(day == 4):
                                cellPeriod = 'H'
                            cellPeriod = cellPeriod+str(period+5)
                            bufferCell = sheet[cellPeriod].value
                            if(bufferCell == ''):
                                sheet[cellPeriod] = labCode
                                sheet[cellPeriod].font = Font(bold=True)
                            else:
                                sheet[cellPeriod].fill = redFill
                                sheet[cellPeriod].font = Font(bold=True)
                                sheet[cellPeriod] = bufferCell+", "+labCode

            for code in completeData:
                actualTeacher = completeData[code]['TH']['teacher'][0:20]
                if(teacherName == actualTeacher):

                    sheet[cell] = code
                    sheet[cellTeacher] = completeData[code]['TH']['teacher']
                    cell = cell[0]+str(int(cell[1:])+1)
                    cellTeacher = cellTeacher[0]+str(int(cellTeacher[1:])+1)
                    if(int(cell[1:])+1 == 24):
                        cell = 'K5'
                        cellTeacher = 'L5'

                    periodos = completeData[code]['TH']['periodos']
                    #print "PERIODOS:"
                    #print periodos
                    for P in range(len(periodos)):
                        period = periodos[P][0]
                        if(period >= 4):
                            period = period +1
                        day = periodos[P][1]
                        if(day == 0):
                            cellPeriod = 'D'
                        elif(day == 1):
                            cellPeriod = 'E'
                        elif(day == 2):
                            cellPeriod = 'F'
                        elif(day == 3):
                            cellPeriod = 'G'
                        elif(day == 4):
                            cellPeriod = 'H'
                        cellPeriod = cellPeriod+str(period+5)
                        bufferCell = sheet[cellPeriod].value
                        if(bufferCell == ''):
                            sheet[cellPeriod] = labCode
                            sheet[cellPeriod].font = Font(bold=True)
                        else:
                            sheet[cellPeriod].fill = redFill
                            sheet[cellPeriod].font = Font(bold=True)
                            sheet[cellPeriod] = bufferCell+", "+code

            #Coloring forbiddenTime

            #Filling all yellow

            for P in range(17):
                for day in range(5):
                    if(P >= 4):
                        period = P +1
                    else:
                        period = P
                    if(day == 0):
                        cellPeriod = 'D'
                    elif(day == 1):
                        cellPeriod = 'E'
                    elif(day == 2):
                        cellPeriod = 'F'
                    elif(day == 3):
                        cellPeriod = 'G'
                    elif(day == 4):
                        cellPeriod = 'H'
                    cellPeriod = cellPeriod+str(period+5)
                    sheet[cellPeriod].fill = yellowFill
                    sheet[cellPeriod].font = Font(bold=True)

            #getting the teacher
            real_Teacher_Name = ''
            for teacherRN in allTeachers:
                if(teacherName == teacherRN[0:20]):
                    real_Teacher_Name = teacherRN

            #Filling white workTime
            if(real_Teacher_Name != ''):
                periodos = allTeachers[real_Teacher_Name].prettyWorkTime
                #print "PERIODOS:"
                #print periodos
                for P in range(len(periodos)):
                    period = periodos[P][0]
                    if(period >= 4):
                        period = period +1
                    day = periodos[P][1]
                    if(day == 0):
                        cellPeriod = 'D'
                    elif(day == 1):
                        cellPeriod = 'E'
                    elif(day == 2):
                        cellPeriod = 'F'
                    elif(day == 3):
                        cellPeriod = 'G'
                    elif(day == 4):
                        cellPeriod = 'H'
                    cellPeriod = cellPeriod+str(period+5)
                    sheet[cellPeriod].fill = whiteFill
                    sheet[cellPeriod].font = Font(bold=True)

        wb.save(filepath)
